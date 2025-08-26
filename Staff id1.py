from openpyxl import load_workbook
from PIL import Image as PILImage, ImageDraw, ImageFont
import io, qrcode
from pathlib import Path
import re

excel_file = "/workspaces/DUMP/EMPLOYEE PROFILE 2025 copy.xlsx"
output_dir = Path("/workspaces/DUMP/Generated_IDs")
output_dir.mkdir(exist_ok=True)

def extract_images(ws, save_dir):
    """Extract embedded images and map them to anchor row/col."""
    photo_map = {}
    for idx, image in enumerate(ws._images, start=1):
        col, row = image.anchor._from.col, image.anchor._from.row
        img_bytes = image._data()
        img = PILImage.open(io.BytesIO(img_bytes))
        filename = save_dir / f"photo_{idx}.png"
        img.save(filename)
        photo_map[(col, row)] = filename
    return photo_map

def create_qr_code(data, save_path):
    qr = qrcode.QRCode(box_size=6, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(save_path)
    return save_path

def create_id_card(emp, photo_path, qr_path, save_path):
    card_w, card_h = 600, 400
    card = PILImage.new("RGB", (card_w, card_h), "white")
    draw = ImageDraw.Draw(card)

    # Staff photo
    if photo_path and Path(photo_path).exists():
        staff_img = PILImage.open(photo_path).resize((160, 160))
        card.paste(staff_img, (30, 100))

    # QR code
    qr_img = PILImage.open(qr_path).resize((140, 140))
    card.paste(qr_img, (420, 120))

    # Text
    try:
        font = ImageFont.truetype("arial.ttf", 18)
    except:
        font = ImageFont.load_default()

    draw.text((220, 40), f"Name: {emp['name']}", font=font, fill="black")
    draw.text((220, 80), f"ID No.: {emp['id']}", font=font, fill="black")
    draw.text((220, 120), f"Position: {emp['position']}", font=font, fill="black")
    draw.text((220, 160), f"DOB: {emp['dob']}", font=font, fill="black")
    draw.text((220, 200), f"Country: {emp['country']}", font=font, fill="black")

    card.save(save_path)

# ---- MAIN PROCESS ----
wb = load_workbook(excel_file)
ws = wb.active

photos = extract_images(ws, output_dir)

# Loop through rows in steps of ~12 (block size)
for row in range(2, 600, 12):  # adjust 600 if more rows
    for col, detail_col in [(2, 10), (18, 25)]:  # (photo col, details col) => (B vs R, J vs Z)
        emp = {
            "name": ws.cell(row=row, column=detail_col).value,
            "id": ws.cell(row=row+1, column=detail_col).value,
            "position": ws.cell(row=row+2, column=detail_col).value,
            "dob": ws.cell(row=row+3, column=detail_col).value,
            "country": ws.cell(row=row+4, column=detail_col).value,
        }
        if not emp["id"]:
            continue  # skip empty slots

        # Match photo by nearest anchor
        photo_path = None
        for (c, r), p_path in photos.items():
            if r in range(row, row+12) and c in range(col, col+6):  # within block
                photo_path = p_path
                break

        # Generate QR code
        qr_filename = output_dir / f"QR_{emp['id']}.png"
        create_qr_code(f"ID:{emp['id']}", qr_filename)

        # Create ID card
        safe_name = re.sub(r'[\\/*?:"<>|]', "", str(emp['name']))
        id_card_path = output_dir / f"ID_{safe_name}_{emp['id']}.png"
        create_id_card(emp, photo_path, qr_filename, id_card_path)

print("✅ All ID cards generated!")
