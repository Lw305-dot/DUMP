import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import re
from datetime import date

master_path = Path("/workspaces/DUMP/Training Progress Tracker.xlsx")
output_dir = Path("/workspaces/DUMP/Employee_Reports")
output_dir.mkdir(exist_ok=True)

def clean_columns(df):
    df.columns = df.columns.str.strip()  # remove leading/trailing spaces
    return df

# trainings = [pd.read_excel(master_path, sheet_name='Cargo Trainings', skiprows=2),
#              pd.read_excel(master_path, sheet_name='DFW Trainings', skiprows=2),
#              pd.read_excel(master_path, sheet_name='AMH Trainings', skiprows=2),
#              pd.read_excel(master_path, sheet_name='CBF Trainings', skiprows=2)]
trainings = [clean_columns(pd.read_excel(master_path, sheet_name=s, skiprows=2))
             for s in ['Cargo Trainings', 'DFW Trainings', 'AMH Trainings', 'CBF Trainings']]
sops  = clean_columns(pd.read_excel(master_path, sheet_name="SOPS", skiprows=2))
exams = clean_columns(pd.read_excel(master_path, sheet_name="EXAMS", skiprows=2))
# sops   = pd.read_excel(master_path, sheet_name="SOPS", skiprows=2)
# exams  = pd.read_excel(master_path, sheet_name="EXAMS", skiprows=2)
# Combine them all in one list
all_dfs = trainings + [sops, exams]

sheet_names = [
    "Cargo Trainings",
    "DFW Trainings",
    "AMH Trainings",
    "CBF Trainings",
    "SOPS",
    "EXAMS",
]

# --- HELPER: CLEAN FILE NAMES ---
def clean_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "_", str(name))

employee_ids = set()

for df in all_dfs:
    if "Emp.No." not in df.columns or "Employee Name" not in df.columns:
        raise ValueError("Sheets must have 'Emp. No.' and 'Employee Name' columns.")
    employee_ids.update(zip(df["Emp. No."], df["Employee Name"]))

employee_list = list(employee_ids)[:20]  # Only first 20 for testing

for emp_id, emp_name in employee_list:
    emp_name_clean = clean_filename(emp_name)
    file_name = f"trainings_for_{emp_name_clean}_{emp_id}.xlsx"
    report_path = output_dir / file_name

    wb = Workbook()
    wb.remove(wb.active)

    for df, sheet_name in zip(all_dfs, sheet_names):
        ws = wb.create_sheet(title=sheet_name)

        emp_df = df[df["Emp. No."] == emp_id]
        if emp_df.empty:
            ws.append([f"No records found for {emp_name} in {sheet_name}"])
            continue
        for r in dataframe_to_rows(emp_df, index=False, header=True):
            ws.append(r)

    wb.save(report_path)

print(f"First 20 employee training files created in: {output_dir}")







# for emp_id, emp_name in employee_ids:
#     emp_name_clean = clean_filename(emp_name)
#     file_name = f"trainings_for_{emp_name_clean}_{emp_id}.xlsx"
#     report_path = output_dir / file_name

#     wb = Workbook()
#     wb.remove(wb.active)

#     for df, sheet_name in zip(all_dfs, sheet_names):
#         ws = wb.create_sheet(title=sheet_name)

#         emp_df = df[df["Emp. No."] == emp_id]
#         if emp_df.empty:
#             ws.append([f"No records found for {emp_name} in {sheet_name}"])
#             continue
#         for r in dataframe_to_rows(emp_df, index=False, header=True):
#             ws.append(r)
#     wb.save(report_path)

# print(f"Employee training files created in: {output_dir}")







# for name, df in zip(sheet_names, all_dfs):
#     date_col = next(
#     (str(c) for c in df.columns if str(c).strip().lower() == "date"),
#     None
# )

#     if date_col:
#         df[date_col] = pd.to_datetime(df[date_col], errors="coerce").dt.strftime("%Y-%m-%d")
#     else:
#         print(f"\n No 'Date' column found in sheet: {name}")

#     print(f"\n--- First 5 rows of '{name}' ---")
#     print(df.head(5))
