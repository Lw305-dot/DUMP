import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime, timedelta
import qrcode
import re

master_path = Path("/workspaces/DUMP/Training Progress Tracker.xlsx")
output_dir = Path("/workspaces/DUMP/Employee_Reports10")
qr_dir = Path("/workspaces/DUMP/Employee_Reports10/QR_Codes")
output_dir.mkdir(exist_ok=True)
qr_dir.mkdir(exist_ok=True) 

BASE_URL = "https://lodigeindustries-my.sharepoint.com/:f:/r/personal/l_karuru_lodige_com/Documents/test"

def find_header_row(sheet_name, file_path):
    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
    for i, row in preview.iterrows():
        if "Emp. No." in row.values and "Employee Name" in row.values:
            return i
    raise ValueError(f"Could not find header row in sheet: {sheet_name}")

def deduplicate_columns(cols):
    seen = {}
    new_cols = []
    for col in cols:
        if col not in seen:
            seen[col] = 0
            new_cols.append(col)
        else:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
    return new_cols

def shorten_sheet_name(empno, empname):
    parts = empname.split()
    second = parts[0] if len(parts) >= 1 else ""
    return f"{empno}_{second}"[:20]

def truncate_title(title):
    if not isinstance(title, str):
        return title
    return title[:20]

def is_marks(value):
    if pd.isna(value): 
        return False
    if isinstance(value, (int, float)): 
        return True
    if isinstance(value, str) and re.match(r"^\d+(\.\d+)?%?$", value.strip()):
        return True
    return False

sheet_names = ["Cargo Trainings", "DFW Trainings", "AMH Trainings", "CBF Trainings", "SOPS", "EXAMS"]
all_dfs = {}
for sheet in sheet_names:
    header_row = find_header_row(sheet, master_path)
    df = pd.read_excel(master_path, sheet_name=sheet, header=header_row)

    df.columns = pd.Series(df.columns).astype(str).str.strip().str.lower().str.replace(r'[^\w\s]', '', regex=True)
    df.columns = deduplicate_columns(df.columns)

    date_cols = [c for c in df.columns if "date" in c.lower()]
    for col in set(date_cols):
        df[col] = pd.to_datetime(df[col], errors="coerce").apply(
            lambda x: x.strftime("%Y-%m-%d") if pd.notnull(x) else ""
        )

    all_dfs[sheet] = df
employee_frames = []

for sheet_df in all_dfs.values():
    if 'emp no' in sheet_df.columns and 'employee name' in sheet_df.columns:
        employee_frames.append(sheet_df[['emp no', 'employee name']])

employees = pd.concat(employee_frames, ignore_index=True).drop_duplicates()
today = datetime.today()
for _, emp in employees.iterrows():
    emp_no = emp['emp no']
    emp_name = emp['employee name']

    # --- Training Dashboard ---
    training_records = []
    for sheet_name, df in all_dfs.items():
        if sheet_name == "EXAMS":
            continue  # Exams handled separately

        rows = df[df['emp no'] == emp_no]
        for _, row in rows.iterrows():
            for date_col in [c for c in df.columns if "date" in c.lower()]:
                date_index = list(df.columns).index(date_col)
                training_name = df.columns[date_index + 1] if date_index + 1 < len(df.columns) else ""
                training_date = pd.to_datetime(row.get(date_col, None), errors='coerce')
                expiry_date = training_date + timedelta(days=365 if sheet_name.lower() == "sops" else 730) if pd.notna(training_date) else None
                days_left = (expiry_date - today).days if expiry_date else None
                status = "VALID" if days_left is not None and days_left >= 0 else (
                    "NOT VALID" if days_left is not None else "NO EXPIRY DATE"
                )

                if training_name:
                    training_records.append([
                        None,
                        training_name,
                        training_date.strftime('%d-%b-%Y') if pd.notna(training_date) else '',
                        expiry_date.strftime('%d-%b-%Y') if pd.notna(expiry_date) else '',
                        days_left,
                        today.strftime('%d-%b-%Y'),
                        status
                    ])

    training_df = pd.DataFrame(training_records, columns=[
        "SN", "TRAININGS", "TRAINING DATE", "EXPIRY DATE", "PERIOD TO EXPIRE", "CURRENT DATE", "STATUS"
    ])
    training_df["SN"] = range(1, len(training_df) + 1)

    # --- Exam Dashboard (Dynamic Version) ---
    exam_df = pd.DataFrame(columns=["SN", "EXAM", "EXAM DATE", "MARKS ATTAINED"])
    if "EXAMS" in all_dfs:
        emp_exam_rows = all_dfs["EXAMS"][all_dfs["EXAMS"]['emp no'] == emp_no]
        exam_records = []

        for _, row in emp_exam_rows.iterrows():
            for idx, col in enumerate(emp_exam_rows.columns):
                if col.startswith("date") and idx + 1 < len(emp_exam_rows.columns):
                    exam_name_col = emp_exam_rows.columns[idx + 1]
                    exam_name_display = exam_name_col.replace("_", " ").title()

                    exam_date = pd.to_datetime(row[col], errors="coerce")
                    mark = row[exam_name_col]

                    if pd.notna(mark) and str(mark).strip() != "":
                        exam_records.append([
                            None,
                            exam_name_display,
                            exam_date.strftime('%d-%b-%Y') if pd.notna(exam_date) else '',
                            mark
                        ])

        if exam_records:
            exam_df = pd.DataFrame(exam_records, columns=["SN", "EXAM", "EXAM DATE", "MARKS ATTAINED"])
            exam_df["SN"] = range(1, len(exam_df) + 1)

            try:
                numeric_marks = exam_df["MARKS ATTAINED"].astype(str).str.replace('%', '').astype(float)
                avg = numeric_marks.mean()
                exam_df.loc["TOTAL"] = ["", "", "TOTAL AVERAGE", f"{avg:.2f}%"]
            except:
                pass

    # --- Write to Excel ---
    wb = Workbook()
    wb.remove(wb.active)

    ws_train = wb.create_sheet(title="Training Dashboard")
    ws_train.append([f"TRAINING DASHBOARD FOR {emp_name} ({emp_no})"])
    for r in dataframe_to_rows(training_df, index=False, header=True):
        ws_train.append(r)

    ws_exam = wb.create_sheet(title="Exam Dashboard")
    ws_exam.append([f"EXAM DASHBOARD FOR {emp_name} ({emp_no})"])
    for r in dataframe_to_rows(exam_df, index=False, header=True):
        ws_exam.append(r)

    file_path = output_dir / f"{emp_name} {emp_no}.xlsx"
    wb.save(file_path)
    # --- Create QR Code for this employee file ---
    file_url = BASE_URL + file_path.name
    qr_file_path = qr_dir / f"Qr_code_for_{emp_name}_{emp_no}.png"
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(file_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(qr_file_path)

print("Dashboards & QR codes created successfully!")

