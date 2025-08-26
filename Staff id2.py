from pathlib import Path
import re
from openpyxl import load_workbook
from PIL import Image as PILImage
import io

# --- CONFIG ---
EXCEL_PATH = Path("/workspaces/DUMP/EMPLOYEE PROFILE 2025 copy.xlsx")  # <-- change
OUTPUT_DIR = Path("Generated_IDs2")         # <-- change
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Columns (1-based). Your structure: photos in B:F or R:V; IDs in J or Z
LEFT_PHOTO_COLS = range(2, 7)    # B..F
RIGHT_PHOTO_COLS = range(18, 23) # R..V
LEFT_ID_COL = 10                 # J
RIGHT_ID_COL = 26                # Z

# Each card repeats downward ~12 rows. We’ll search within this height window.
BLOCK_HEIGHT = 12

# Match things like Q0011, Q0048, etc. Loosen if your IDs vary.
ID_REGEX = re.compile(r"^[A-Za-z]{0,3}\d{3,}$")  # e.g., Q0011, EMP1234, 0007


def image_anchor_top_left(img):
    """
    Get 1-based (col, row) for the image's top-left anchor.
    openpyxl stores these as 0-based; convert to 1-based for worksheet.cell().
    """
    # Works for ClientAnchor/OneCellAnchor
    anchor_from = getattr(img.anchor, "_from", None)
    if anchor_from is None:
        # Fallback for other anchor types
        # Try 'img.anchor.from' if available (older openpyxl)
        anchor_from = getattr(img.anchor, "from", None)
    col = anchor_from.col + 1
    row = anchor_from.row + 1
    return col, row


def find_id_in_block(ws, start_row, id_col, height=BLOCK_HEIGHT):
    """
    Look down from start_row for the first non-empty 'ID NO.' value in id_col
    that looks like an ID by regex.
    """
    end_row = start_row + height - 1
    max_row = min(end_row, ws.max_row)
    for r in range(start_row, max_row + 1):
        val = ws.cell(row=r, column=id_col).value
        if val is None:
            continue
        text = str(val).strip()
        # Ignore the "ID NO.:" label cell if it appears
        if text.upper().startswith("ID NO"):
            continue
        if ID_REGEX.match(text):
            return text
    return None


def save_photo_as_pdf(img_bytes, pdf_path):
    """
    Convert the embedded image bytes to a single-page PDF with no extra text.
    """
    with PILImage.open(io.BytesIO(img_bytes)) as im:
        # Ensure RGB and a reasonable size; keep original size
        rgb = im.convert("RGB")
        rgb.save(pdf_path, "PDF")


def run():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # openpyxl keeps embedded images on the worksheet
    images = getattr(ws, "_images", [])
    if not images:
        print("No embedded images found.")
        return

    seen_ids = set()
    made = 0

    for img in images:
        # 1) Where is the photo anchored?
        col, row = image_anchor_top_left(img)

        # 2) Decide which ID column to search (left or right card)
        if col in LEFT_PHOTO_COLS:
            id_col = LEFT_ID_COL
        elif col in RIGHT_PHOTO_COLS:
            id_col = RIGHT_ID_COL
        else:
            # Not inside a recognized photo block; skip
            continue

        # 3) Find the staff ID in the nearby block rows
        staff_id = find_id_in_block(ws, row, id_col, BLOCK_HEIGHT)
        if not staff_id:
            # Try nudging the start row up a little if photos start a row below text
            staff_id = find_id_in_block(ws, max(1, row - 2), id_col, BLOCK_HEIGHT + 2)

        if not staff_id:
            # Still no match; skip this image
            continue

        # 4) Save the photo as a PDF named by the ID
        safe_id = re.sub(r'[\\/*?:"<>|]', "", staff_id)
        pdf_path = OUTPUT_DIR / f"{safe_id}.pdf"

        # Get raw bytes from the embedded image
        # openpyxl Image stores the raw data accessible via _data()
        img_bytes = img._data()
        save_photo_as_pdf(img_bytes, pdf_path)

        if safe_id not in seen_ids:
            seen_ids.add(safe_id)
            made += 1

    print(f"Done. Created {made} PDF(s) in: {OUTPUT_DIR}")


if __name__ == "__main__":
    run()
