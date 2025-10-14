import pandas as pd 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font,PatternFill
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime, timedelta
import re
import qrcode
from PIL import Image as PILImage, ImageDraw, ImageFont

PHOTO_DIR = Path("/workspaces/DUMP/Generated_IDs3")
master_path = Path("/workspaces/DUMP/Training Progress Tracker.xlsx")
training_list_path = Path("/workspaces/DUMP/MASTER LIST Module number.xlsx")

output_dir = Path("/workspaces/DUMP/Employee_Reports31")
# trainingid_dir= Path("/workspaces/DUMP/Employee_Reports30/TrainingIDs")
# qr_dir = Path("/workspaces/DUMP/Employee_Reports30/QR_Codes")
output_dir.mkdir(exist_ok=True)
# qr_dir.mkdir(exist_ok=True)
# trainingid_dir.mkdir(exist_ok=True)

BASE_URL = "https://lodigeindustries-my.sharepoint.com/:f:/r/personal/l_karuru_lodige_com/Documents/Trainings/"
logo_path=Path("/workspaces/DUMP/logo.png")
training_sheets = pd.read_excel(training_list_path, sheet_name=["Manuals", "SOPs"])
training_lookup_df = pd.concat(training_sheets.values(), ignore_index=True)
logo_path="/workspaces/DUMP/logo.png"

def shorten_name(emp_name, max_length=20):
    """
    Take up to 3 words of the name.
    Count characters only (ignore spaces).
    If total length > max_length, cut the 3rd word and add '...'.
    """
    parts = emp_name.split()
    if not parts:
        return ""

    # Take up to 3 words
    parts = parts[:3]

    # Calculate total length without spaces
    total_len = sum(len(p) for p in parts)

    if total_len <= max_length:
        return " ".join(parts)

    # If too long and 3rd word exists → trim it
    if len(parts) == 3:
        first_two_len = len(parts[0]) + len(parts[1])
        allowed_third_len = max_length - first_two_len
        if allowed_third_len > 0:
            parts[2] = parts[2][:allowed_third_len] + "..."
            return " ".join(parts)
        else:
            # Not enough space for third word at all
            return f"{parts[0]} {parts[1]}..."
    else:
        # Only 1–2 words, just truncate whole name
        return emp_name[:max_length] + "..."


# def create_id_card(emp_no, emp_name, qr_file_path, logo_path):
#     # Constants (ID card size)
#     CARD_WIDTH, CARD_HEIGHT = 346, 210
#     MARGIN = 10
#     GAP = 8
#     IMAGE_SIZE = 100

#     # Load images
#     photo_path = PHOTO_DIR / f"{emp_no}.png"
#     if not photo_path.exists():
#         print(f"No photo for {emp_no}, skipping ID card.")
#         return

#     photo_img = PILImage.open(photo_path).convert("RGB")
#     qr_img = PILImage.open(qr_file_path).convert("RGB")
#     logo_img = PILImage.open(logo_path).convert("RGBA")

#     # Resize photo & QR (same size)
#     photo_img = photo_img.resize((IMAGE_SIZE, IMAGE_SIZE))
#     qr_img = qr_img.resize((IMAGE_SIZE, IMAGE_SIZE))

#     # Fonts
#     try:
#         title_font = ImageFont.truetype("DejaVuSans-Bold.ttf", 18)
#         label_font = ImageFont.truetype("DejaVuSans.ttf", 12)
#         footer_font = ImageFont.truetype("DejaVuSans.ttf", 9)
#     except IOError:
#         title_font = ImageFont.load_default()
#         label_font = ImageFont.load_default()
#         footer_font = ImageFont.load_default()

#     # Create blank card
#     card = PILImage.new("RGB", (CARD_WIDTH, CARD_HEIGHT), "white")
#     draw = ImageDraw.Draw(card)

#     # Border
#     draw.rectangle([0, 0, CARD_WIDTH - 1, CARD_HEIGHT - 1], outline="black", width=1)

#     # === Prepare text block ===
#     emp_name_short = shorten_name(emp_name, max_length=20)
#     title_text = "TRAINING ID"
#     staff_text = f"Staff ID : {emp_no}"
#     name_text = f"Name : {emp_name_short}"

#     # Dummy draw for measuring text
#     dummy_img = PILImage.new("RGB", (1, 1))
#     dummy_draw = ImageDraw.Draw(dummy_img)

#     # Measure text heights
#     title_bbox = dummy_draw.textbbox((0, 0), title_text, font=title_font)
#     staff_bbox = dummy_draw.textbbox((0, 0), staff_text, font=label_font)
#     name_bbox = dummy_draw.textbbox((0, 0), name_text, font=label_font)

#     title_height = title_bbox[3] - title_bbox[1]
#     staff_height = staff_bbox[3] - staff_bbox[1]
#     name_height = name_bbox[3] - name_bbox[1]

#     # Define spacings
#     spacing_between_lines = 5
#     total_text_block_height = (
#         title_height
#         + spacing_between_lines
#         + staff_height
#         + spacing_between_lines
#         + name_height
#     )

#     # === Resize logo to match total block height ===
#     logo_target_height = total_text_block_height
#     logo_ratio = logo_img.width / logo_img.height
#     logo_width = int(logo_target_height * logo_ratio)
#     logo_img = logo_img.resize((logo_width, logo_target_height))

#     # --- Block heights ---
#     top_section_height = max(logo_img.height, total_text_block_height)
#     middle_block_height = IMAGE_SIZE
#     footer_text = "Quality management compliance, authorized under Training protocols."
#     footer_bbox = dummy_draw.textbbox((0, 0), footer_text, font=footer_font)
#     footer_height = footer_bbox[3] - footer_bbox[1]

#     total_content_height = top_section_height + middle_block_height + footer_height
#     remaining_space = CARD_HEIGHT - total_content_height - 2 * MARGIN
#     if remaining_space < 0:
#         remaining_space = 0
#     gap_y = remaining_space // 2

#     # --- Place top block ---
#     top_y = MARGIN
#     logo_x = MARGIN
#     logo_y = top_y + (top_section_height - logo_img.height) // 2
#     card.paste(logo_img, (logo_x, logo_y), logo_img)

#     text_x = logo_x + logo_img.width + 12
#     text_y = top_y + (top_section_height - total_text_block_height) // 2
#     draw.text((text_x, text_y), title_text, font=title_font, fill="black")
#     draw.text(
#         (text_x, text_y + title_height + spacing_between_lines),
#         staff_text,
#         font=label_font,
#         fill="black",
#     )
#     draw.text(
#         (
#             text_x,
#             text_y
#             + title_height
#             + spacing_between_lines
#             + staff_height
#             + spacing_between_lines,
#         ),
#         name_text,
#         font=label_font,
#         fill="black",
#     )

#     # --- middle block (photo + QR) with equal margins ---
#     images_y = top_y + top_section_height + gap_y
#     num_gaps = 3
#     available_width = CARD_WIDTH - 2 * MARGIN
#     total_images_width = photo_img.width + qr_img.width
#     gap_x = (available_width - total_images_width) // num_gaps

#     # Place photo
#     photo_x = MARGIN + gap_x
#     card.paste(photo_img, (photo_x, images_y))

#     # Place QR
#     qr_x = photo_x + photo_img.width + gap_x
#     card.paste(qr_img, (qr_x, images_y))

#     # --- Place footer ---
#     footer_y = images_y + middle_block_height + gap_y
#     footer_w = footer_bbox[2] - footer_bbox[0]
#     footer_x = (CARD_WIDTH - footer_w) // 2
#     draw.text((footer_x, footer_y), footer_text, font=footer_font, fill="gray")

#     # Save card
#     card.save(trainingid_dir / f"ID_CARD_{emp_no}.png")

    
def find_header_row(sheet_name, file_path):
    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
    for i, row in preview.iterrows():
        if "Emp. No." in row.values and "Employee Name" in row.values:
            return i
    raise ValueError(f"Could not find header row in sheet: {sheet_name}")

def deduplicate_columns(cols):
    seen = {}
    new_cols = []
    for col in cols:
        if col not in seen:
            seen[col] = 0
            new_cols.append(col)
        else:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
    return new_cols
training_lookup_df.columns = deduplicate_columns( training_lookup_df.columns.str.strip
                                                 ().str.lower()
)
training_lookup = {
    str(row["trainings"]).strip().lower(): {
        "code": row["code_1"] if pd.notna(row.get("code_1")) else row.get("code", ""),
        "facility": row["facility"],
        "category": row["category"]
    }
    for _, row in training_lookup_df.iterrows()
}
EXAM_OUTDATED_DAYS = 5000  
WRONGLY_FORMATTED_AFTER_DAYS = -6000 
HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78") 
TITLE_FONT = Font(bold=True, color="FFFFFF")
ROW_RED_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE") 
ROW_YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFFACD", end_color="FFFACD")  # Light Yellow
def style_sheet(ws):
    # Style the first row as header (merged row should be done outside if needed)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    header_cell = ws.cell(row=1, column=1)  # First cell in first row
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.font = Font(bold=True, size=14)
    
    header_map = {}
    for idx, cell in enumerate(ws[2], start=1):
        if cell.value:
            header_map[str(cell.value).strip().upper()] = idx


    # Set alignment for the rest of the sheet
    align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Define thin border
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[2]:
        cell.font = TITLE_FONT
        cell.alignment = align
        cell.fill = HEADER_FILL
        cell.border = thin_border

 
    # Apply styles to all cells except the first row
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = align
            cell.border = thin_border
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD-MMM-YYYY'
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):  # EXAM DATE column
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD-MMM-YYYY'      
    status_col = None
    for idx, cell in enumerate(ws[2], start=1):
        if str(cell.value).strip().upper() == "STATUS":
            status_col = idx
            break
    if status_col:
        for r in range(3, ws.max_row + 1):
            status_val = ws.cell(row=r, column=status_col).value
            status_str = str(status_val).strip().upper()
            if status_str == "NOT VALID"or status_str == "LOW PERCENTAGE":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = ROW_RED_FILL
            elif status_str == "EXPIRING SOON":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = ROW_YELLOW_FILL
                    
    # Format "MARKS ATTAINED" column as percentage if it exists 
    if "MARKS ATTAINED" in header_map:
        marks_col = header_map["MARKS ATTAINED"]
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=marks_col)
            if isinstance(cell.value, (int, float)):
               cell.number_format = "0.00%"
    # Auto-fit column widths, skipping the first row
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[1:]:  # Skip the first row
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)  # add margin
        ws.column_dimensions[col_letter].width = adjusted_width

sheet_names = ["Cargo Trainings", "DFW Trainings", "AMH Trainings", "CBF Trainings", "SOPs","QNL Trainings","Other Trainings","Work Instruction","EXAMS"]
all_dfs = {}
for sheet in sheet_names:
    header_row = find_header_row(sheet, master_path)
    df = pd.read_excel(master_path, sheet_name=sheet, header=header_row)

    cleaned_cols = []
    for col in df.columns:
        col_str = str(col)

        # Strict clean for Emp No and Employee Name
        if col_str.strip() in ["Emp. No.", "Employee Name"]:
            col_str = (
                pd.Series([col_str])
                .astype(str)
                .str.strip()
                .str.lower()
                .str.replace(r"[^\w\s]", "", regex=True)
                .iloc[0]
            )

        # ✅ Also normalize date-like headers (for training + exams to work)
        elif "date" in col_str.lower():
            col_str = (
                pd.Series([col_str])
                .astype(str)
                .str.strip()
                .str.lower()
                .str.replace(r"[^\w\s]", "", regex=True)
                .iloc[0]
            )

        cleaned_cols.append(col_str)


    df.columns = deduplicate_columns(cleaned_cols)
    all_dfs[sheet] = df

cargo_df = all_dfs[sheet_names[0]]
employees = cargo_df[['emp no', 'employee name']].drop_duplicates()
for sheet_name, sheet_df in all_dfs.items():
    if sheet_name == sheet_names[0]:  # skip Cargo (already taken)
        continue
    if 'emp no' in sheet_df.columns and 'employee name' in sheet_df.columns:
        new_emps = sheet_df[['emp no', 'employee name']].drop_duplicates()
        # Add only new employees not already in the list
        employees = pd.concat([employees, new_emps]).drop_duplicates(subset=['emp no'])

today = datetime.today()
# --- Process employees ---
for _, emp in employees.iterrows():
    emp_no = emp['emp no']
    emp_name = emp['employee name']
    emp_info_df = pd.read_excel("Employee Designations.xlsx")
    emp_info_df.columns = emp_info_df.columns.str.lower()
    emp_desg = emp_info_df.loc[emp_info_df["emp. no."] == emp_no]
    emp_desg_value = emp_desg["job description"].iloc[0] if not emp_desg.empty else "N/A"
    
    # --- Collect Training Records ---
    training_records = []
    for sheet_name, df in all_dfs.items():
        if sheet_name.upper() == "EXAMS":
            continue  # exams handled separately
        rows = df[df['emp no'] == emp_no]
        for _, row in rows.iterrows():
            for date_col in [c for c in df.columns if "date" in c.lower()]:
                training_date = pd.to_datetime(row.get(date_col, None), errors="coerce")

                # ✅ Skip if no date
                if pd.isna(training_date):
                    continue

                date_index = list(df.columns).index(date_col)
                training_name = df.columns[date_index + 1] if date_index + 1 < len(df.columns) else ""
               
                expiry_date = training_date + timedelta(days=365 if sheet_name.lower()=="sops" else 730)
                days_left = (expiry_date - today).days if expiry_date else None
                remark = ""
                if days_left is not None and days_left < WRONGLY_FORMATTED_AFTER_DAYS:
                    remark = "this is a wronlgy formatted date and it will be corrected in the future"
                status = (
                    "EXPIRING SOON" if days_left is not None and 1 <= days_left < 40
                    else "VALID" if days_left is not None and days_left >= 40
                    else "NOT VALID" if days_left is not None
                    else "NOT Applicable"
                )
                if training_name:
                    lookup_key = str(training_name).strip().lower()
                    training_info = training_lookup.get(lookup_key, {"code": "", "facility": "", "category": ""})
                    training_records.append([
                        None,
                        f"{training_name} ({sheet_name})",
                        training_info["facility"],
                        training_info["code"],
                        training_info["category"],
                        training_date.strftime('%d-%b-%Y'),
                        expiry_date.strftime('%d-%b-%Y'),
                        days_left,
                        today.strftime('%d-%b-%Y'),
                        status,
                        remark
                    ])

    training_df = pd.DataFrame(training_records, columns=[
        "SN", "TRAININGS","FACILITY","CODE","CATEGORY","TRAINING DATE", "EXPIRY DATE", "PERIOD TO EXPIRE", "LAST UPDATE", "STATUS","REMARKS"
    ])
    training_df.drop_duplicates(subset=["TRAININGS", "TRAINING DATE"], inplace=True)
    training_df["SN"] = range(1, len(training_df) + 1)

    # Add emp_desg_value as a final row in the training dashboard
    # Fill other columns with empty strings for clarity
    if emp_desg_value != "N/A":
        desg_row = ["", f"YOU ARE LISTED AS ; {emp_desg_value}", "& THIS IS YOUR TRAINING DASHBOARD"] + ["" for _ in range(len(training_df.columns)-3)]
        training_df.loc[len(training_df)] = desg_row
    # --- Collect Exam Records ---
    exam_records = []
    if "EXAMS" in all_dfs:
        emp_exam_rows = all_dfs["EXAMS"][all_dfs["EXAMS"]['emp no'] == emp_no]
        for _, row in emp_exam_rows.iterrows():
            for idx, col in enumerate(emp_exam_rows.columns):
                if col.startswith("date") and idx + 1 < len(emp_exam_rows.columns):
                    exam_name_col = emp_exam_rows.columns[idx + 1]
                    exam_name_display = exam_name_col.replace("_", " ").title()
                    exam_date = pd.to_datetime(row[col], errors="coerce")
                    mark = row[exam_name_col]

                    status = "NOT VALID"
                    mark_comment = ""
                    date_comment = ""

                    mark_value = pd.to_numeric(str(mark).replace("%", ""), errors="coerce")
                    if pd.notna(mark_value):
                        if mark_value <= 1:
                            mark_value *= 100
                        mark_value = round(mark_value, 2)
                        if mark_value < 75:
                            status = "low percentage"
                            mark_comment = "This is a low mark, please retake the exam and improve your score."
                        elif mark_value >= 98:
                            status = "perfect score"
                            mark_comment = " Congratulations on achieving a perfect score!!!!"
                        else:
                            status = "VALID"
                            mark_comment = "Approved Score."
                    
                    if pd.isna(exam_date):
                        date_comment = "Kindly redo your exam the exam is not recorded properly"
                        status = "NOT VALID"
                    else:
                        age_days = (today - exam_date).days
                        if age_days > EXAM_OUTDATED_DAYS:
                            date_comment= "Kindly retake your exam this exam is outdated"
                            status = "NOT VALID"
                        else:
                            date_comment = "date is valid"
                    if pd.notna(exam_date) or pd.notna(mark_value):
                       comment = " ".join(filter(None, [mark_comment, date_comment])).strip()
                       exam_records.append([
                             None,
                             exam_name_display,
                             exam_date.strftime('%d-%b-%Y') if pd.notna(exam_date) else '',
                            f"{mark_value:.2f}%" if pd.notna(mark_value) else mark,
                            mark_value, 
                            status,
                            comment
                    ])

    exam_df = pd.DataFrame(exam_records, columns=["SN", "EXAM", "EXAM DATE", "MARKS ATTAINED","MARK VALUE","STATUS","COMMENTS"])
    if not exam_df.empty:
        exam_df.drop_duplicates(subset=["EXAM", "EXAM DATE"], inplace=True) 
        exam_df["SN"] = range(1, len(exam_df) + 1)
        try:
            avg = exam_df["MARK VALUE"].dropna().mean()
            exam_df.loc[len(exam_df)] = ["", "", "TOTAL AVERAGE", avg, ""]
        except:
            pass
    export_df = exam_df.drop(columns=["MARK VALUE"], errors="ignore")
    if not export_df.empty:
        try:
            avg = exam_df["MARK VALUE"].dropna().mean()
            export_df.loc[len(export_df)] = ["", "", "TOTAL AVERAGE", f"{avg:.2f}%", "", ""]
        except:
            pass
    # --- Write ONE Excel file with 2 sheets ---
    wb = Workbook()
    wb.remove(wb.active)

    ws_train = wb.create_sheet(title="Training Dashboard")
    ws_train.append([f"TRAINING DASHBOARD FOR {emp_name} ({emp_no})"])
    for r in dataframe_to_rows(training_df, index=False, header=True):
        ws_train.append(r)
    style_sheet(ws_train)

    ws_exam = wb.create_sheet(title="Exam Dashboard")
    ws_exam.append([f"EXAM DASHBOARD FOR {emp_name} ({emp_no})"])
    for r in dataframe_to_rows(export_df, index=False, header=True):
        ws_exam.append(r)
    style_sheet(ws_exam)

    # safe_emp_name = re.sub(r'[\\/*?:"<>|]', "", str(emp_name)).strip()
    safe_emp_name = str(emp_name).strip()
    if safe_emp_name.lower() == "nan" or safe_emp_name == "" or pd.isna(emp_name):
          # If no name, just use emp_no
       filename = f"{emp_no}.xlsx"
       qr_filename = f"Qr_code_for_{emp_no}.png"

    else:
        safe_emp_name = re.sub(r'[\\/*?:"<>|]', "", safe_emp_name)
        filename = f"{safe_emp_name} {emp_no}.xlsx"
        qr_filename = f"Qr_code_for_{safe_emp_name}_{emp_no}.png"
    file_path = output_dir/filename
    wb.save(file_path)
    
    # file_url = BASE_URL + file_path.name
    # qr_file_path = qr_dir/qr_filename
    # qr = qrcode.QRCode(version=1, box_size=10, border=4)
    # qr.add_data(file_url)
    # qr.make(fit=True)
    # img = qr.make_image(fill_color="black", back_color="white")
    # img.save(qr_file_path)
    # create_id_card(emp_no, emp_name, qr_file_path,logo_path)

print("Training reports created successfully!")
