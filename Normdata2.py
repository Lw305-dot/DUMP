import pandas as pd 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime, timedelta
import re
import qrcode
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import os

master_path = Path("/workspaces/DUMP/Training Progress Tracker.xlsx")
output_dir = Path("/workspaces/DUMP/Employee_Reports12")

qr_dir = Path("/workspaces/DUMP/Employee_Reports12/QR_Codes")
output_dir.mkdir(exist_ok=True)
qr_dir.mkdir(exist_ok=True) 

BASE_URL = "https://drive.google.com/drive/folders/1ghaHhqs2qiCKBtsm8Z20mT-XcY7hehg_?usp=drive_link"
SERVICE_ACCOUNT_FILE = '/workspaces/DUMP/service_account.json'
SCOPES = ['https://www.googleapis.com/auth/drive.file']
MAIN_FOLDER_ID = "1TfaHhqs2qiCKBtsm8Z20mT-XcY7hehg_"   
QR_FOLDER_ID   = "1r0awgcwbUbvSIGsQ_bP_sHIUJyBwVdgV"  
def authenticate_drive():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES
    )
    service = build('drive', 'v3', credentials=creds)
    return service
def list_permissions(folder_id):
    service = authenticate_drive()
    permissions = service.permissions().list(fileId=folder_id).execute()
    print("Folder permissions:", permissions)
def list_files(folder_id):
    service = authenticate_drive()
    results = service.files().list(
        q=f"'{folder_id}' in parents",
        fields="files(id, name)"
    ).execute()
    for f in results.get('files', []):
        print(f)
def upload_to_drive(file_path, folder_id):
    service = authenticate_drive()
    file_metadata = {
        'name': os.path.basename(file_path),
        'parents': [folder_id]
    }
    media = MediaFileUpload(file_path, resumable=True)
    file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    print(f"✅ Uploaded {file_path} → Drive ID: {file.get('id')}")
    return file.get('id')

def find_header_row(sheet_name, file_path):
    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
    for i, row in preview.iterrows():
        if "Emp. No." in row.values and "Employee Name" in row.values:
            return i
    raise ValueError(f"Could not find header row in sheet: {sheet_name}")

def deduplicate_columns(cols):
    seen = {}
    new_cols = []
    for col in cols:
        if col not in seen:
            seen[col] = 0
            new_cols.append(col)
        else:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
    return new_cols

sheet_names = ["Cargo Trainings", "DFW Trainings", "AMH Trainings", "CBF Trainings", "SOPS", "EXAMS"]
all_dfs = {}
for sheet in sheet_names:
    header_row = find_header_row(sheet, master_path)
    df = pd.read_excel(master_path, sheet_name=sheet, header=header_row)

    df.columns = pd.Series(df.columns).astype(str).str.strip().str.lower().str.replace(r'[^\w\s]', '', regex=True)
    df.columns = deduplicate_columns(df.columns)

    date_cols = [c for c in df.columns if "date" in c.lower()]
    for col in set(date_cols):
        df[col] = pd.to_datetime(df[col], errors="coerce").apply(
            lambda x: x.strftime("%Y-%m-%d") if pd.notnull(x) else ""
        )

    all_dfs[sheet] = df
    # if sheet.upper() == "EXAMS":
    #     print("\n📑 Raw EXAMS DataFrame loaded from Excel:")
    #     print(df.head(20))
    #     print("\n🔎 Columns:", df.columns.tolist())
    #     print("Shape:", df.shape)
cargo_df = all_dfs[sheet_names[0]]
employees = cargo_df[['emp no', 'employee name']].drop_duplicates()
for sheet_name, sheet_df in all_dfs.items():
    if sheet_name == sheet_names[0]:  # skip Cargo (already taken)
        continue
    if 'emp no' in sheet_df.columns and 'employee name' in sheet_df.columns:
        new_emps = sheet_df[['emp no', 'employee name']].drop_duplicates()
        # Add only new employees not already in the list
        employees = pd.concat([employees, new_emps]).drop_duplicates(subset=['emp no'])

today = datetime.today()
# --- Process employees ---
for _, emp in employees.iterrows():
    emp_no = emp['emp no']
    emp_name = emp['employee name']

    # --- Collect Training Records ---
    training_records = []
    for sheet_name, df in all_dfs.items():
        if sheet_name.upper() == "EXAMS":
            continue  # exams handled separately
        rows = df[df['emp no'] == emp_no]
        for _, row in rows.iterrows():
            for date_col in [c for c in df.columns if "date" in c.lower()]:
                training_date = pd.to_datetime(row.get(date_col, None), errors="coerce")

                # ✅ Skip if no date
                if pd.isna(training_date):
                    continue

                date_index = list(df.columns).index(date_col)
                training_name = df.columns[date_index + 1] if date_index + 1 < len(df.columns) else ""

                expiry_date = training_date + timedelta(days=365 if sheet_name.lower()=="sops" else 730)
                days_left = (expiry_date - today).days if expiry_date else None
                status = (
                    "VALID" if days_left is not None and days_left >= 0
                    else "NOT VALID" if days_left is not None
                    else "NOT Applicable"
                )
                if training_name:
                    training_records.append([
                        None,
                        f"{training_name} ({sheet_name})", 
                        training_date.strftime('%d-%b-%Y'),
                        expiry_date.strftime('%d-%b-%Y'),
                        days_left,
                        today.strftime('%d-%b-%Y'),
                        status
                    ])

    training_df = pd.DataFrame(training_records, columns=[
        "SN", "TRAININGS", "TRAINING DATE", "EXPIRY DATE", "PERIOD TO EXPIRE", "CURRENT DATE", "STATUS"
    ])
    training_df.drop_duplicates(subset=["TRAININGS", "TRAINING DATE"], inplace=True)
    training_df["SN"] = range(1, len(training_df) + 1)

    # --- Collect Exam Records ---
    exam_records = []
    if "EXAMS" in all_dfs:
        emp_exam_rows = all_dfs["EXAMS"][all_dfs["EXAMS"]['emp no'] == emp_no]
        for _, row in emp_exam_rows.iterrows():
            for idx, col in enumerate(emp_exam_rows.columns):
                if col.startswith("date") and idx + 1 < len(emp_exam_rows.columns):
                    exam_name_col = emp_exam_rows.columns[idx + 1]
                    exam_name_display = exam_name_col.replace("_", " ").title()
                    exam_date = pd.to_datetime(row[col], errors="coerce")
                    mark = row[exam_name_col]
                    if pd.notna(mark) and str(mark).strip() != "":
                        exam_records.append([
                            None,
                            exam_name_display,
                            exam_date.strftime('%d-%b-%Y') if pd.notna(exam_date) else '',
                            mark
                        ])

    exam_df = pd.DataFrame(exam_records, columns=["SN", "EXAM", "EXAM DATE", "MARKS ATTAINED"])
    # print("\n📘 Cleaned Exam DataFrame:")
    # print(exam_df.head(20))
    # print("\n🔎 Columns:", exam_df.columns.tolist())
    if not exam_df.empty:
        exam_df.drop_duplicates(subset=["EXAM", "EXAM DATE"], inplace=True) 
        exam_df["SN"] = range(1, len(exam_df) + 1)
        try:
            numeric_marks = exam_df["MARKS ATTAINED"].astype(str).str.replace('%', '').astype(float)
            avg = numeric_marks.mean()
            exam_df.loc["TOTAL"] = ["", "", "TOTAL AVERAGE", f"{avg:.2f}%"]
        except:
            pass

    # --- Write ONE Excel file with 2 sheets ---
    wb = Workbook()
    wb.remove(wb.active)

    ws_train = wb.create_sheet(title="Training Dashboard")
    ws_train.append([f"TRAINING DASHBOARD FOR {emp_name} ({emp_no})"])
    for r in dataframe_to_rows(training_df, index=False, header=True):
        ws_train.append(r)

    ws_exam = wb.create_sheet(title="Exam Dashboard")
    ws_exam.append([f"EXAM DASHBOARD FOR {emp_name} ({emp_no})"])
    for r in dataframe_to_rows(exam_df, index=False, header=True):
        ws_exam.append(r)

    safe_emp_name = re.sub(r'[\\/*?:"<>|]', "", str(emp_name)).strip()
    file_path = output_dir / f"{safe_emp_name} {emp_no}.xlsx"
    wb.save(file_path)
    # Upload to Drive
    file_url = upload_to_drive(file_path, MAIN_FOLDER_ID)
    # file_url = BASE_URL + file_path.name
    qr_file_path = qr_dir / f"Qr_code_for_{safe_emp_name}_{emp_no}.png"
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(file_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(qr_file_path)

    qr_url = upload_to_drive(qr_file_path, QR_FOLDER_ID)
    
print("✅ Dashboards & QR codes uploaded successfully!")
print("Excel URL:", file_url)
print("QR URL:", qr_url)
print("Dashboards & QR codes created successfully!")
