import os
import re
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# === CONFIG ===
source_folder = Path("/workspaces/DUMP/Employee_Reports35")
output_dir = Path("/workspaces/DUMP/Expired Trainings")
output_dir.mkdir(exist_ok=True)
output_file = output_dir / "expired_training_summary2.xlsx"

expired_data = []
sop_data = []

KEYWORDS = ["NOT VALID", "EXPIRING SOON"]

for file in os.listdir(source_folder):
    if not file.endswith(".xlsx"):
        continue

    file_path = source_folder / file
    wb = load_workbook(file_path, data_only=True)

    # --- Check for the target sheet ---
    if "Training Dashboard" not in wb.sheetnames:
        print(f"⚠️ Skipping {file} (no 'Training Dashboard' sheet)")
        continue

    ws = wb["Training Dashboard"]

    # --- Extract name and staff number from header or filename ---
    name = None
    staff_no = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for value in row:
            if value and isinstance(value, str):
                v = value.strip()
                if "q" in v.lower() and any(ch.isdigit() for ch in v):
                    staff_no = v
                elif "name" in v.lower():
                    name = v
    if not name:
        name = Path(file).stem
    if not staff_no:
        staff_no = "Unknown"

    # --- Convert worksheet to DataFrame ---
    df = pd.DataFrame(ws.values)

    # Find the header row (first row that contains "TRAINING" or "STATUS")
    header_row_index = None
    for i, row in df.iterrows():
        if any(isinstance(v, str) and "trainings" in v.lower() for v in row):
            header_row_index = i
            break

    if header_row_index is None:
        print(f"⚠️ Could not find header row in {file}")
        continue

    df.columns = df.iloc[header_row_index]
    df = df.drop(range(0, header_row_index + 1)).reset_index(drop=True)
    df = df.apply(lambda x: x.map(lambda v: str(v).strip().upper() if isinstance(v, str) else v))

    # --- Check "STATUS" column for NOT VALID / EXPIRING SOON ---
    if "STATUS" not in df.columns:
        print(f"⚠️ No 'STATUS' column in {file}")
        continue

    invalid_rows = df[df["STATUS"].isin(KEYWORDS)]
    if invalid_rows.empty:
        print(f"ℹ️ No expiring or invalid trainings found in {file}")
        continue

    # --- Separate EQUIPMENTS vs SOPs ---
    for _, row in invalid_rows.iterrows():
        training_name = str(row.get("TRAININGS", "Unknown Training"))
        status = str(row.get("STATUS", "Unknown"))

        match = re.findall(r"\(([^()]*)\)", training_name)
        category_tag = match[-1].lower() if match else ""
        record = {
            "Name": name,
            "Staff number": staff_no,
            "Training name": training_name,
            "Status": status,
        }
        if "equip" in category_tag:
            expired_data.append(record)
        elif "sop" in category_tag or "other" in category_tag:
            sop_data.append(record)
        else:
            sop_data.append(record)

def group_records(records):
    grouped = {}
    for rec in records:
        key = (rec["Staff number"], rec["Name"])
        grouped.setdefault(key, [])
        grouped[key].append((rec["Training name"], rec["Status"]))

    # flatten into one row per person
    combined = []
    for (staff_no, name), trainings in grouped.items():
        row = {"Staff number": staff_no, "Name": name}
        for i, (tname, status) in enumerate(trainings, start=1):
            row[f"Training {i}"] = tname
            row[f"Status {i}"] = status
        combined.append(row)
    return combined
# === SAVE GROUPED RESULTS ===
if not expired_data and not sop_data:
    print("⚠️ No matches found in any file.")
else:
    with pd.ExcelWriter(output_file) as writer:
        if expired_data:
            pd.DataFrame(group_records(expired_data)).to_excel(writer, sheet_name="Equipments", index=False)
        if sop_data:
            pd.DataFrame(group_records(sop_data)).to_excel(writer, sheet_name="SOPs & Other Trainings", index=False)
    print(f"✅ Summary file created successfully: {output_file}")