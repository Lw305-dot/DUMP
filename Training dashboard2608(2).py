import pandas as pd 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font,numbers
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime, timedelta
import re
import qrcode
from PIL import Image as PILImage, ImageDraw, ImageFont

PHOTO_DIR = Path("/workspaces/DUMP/Generated_IDs3")
master_path = Path("/workspaces/DUMP/Training Progress Tracker.xlsx")
training_list_path = Path("/workspaces/DUMP/MASTER LIST Module number.xlsx")

output_dir = Path("/workspaces/DUMP/Employee_Reports21")
trainingid_dir= Path("/workspaces/DUMP/Employee_Reports21/TrainingIDs")
qr_dir = Path("/workspaces/DUMP/Employee_Reports21/QR_Codes")
output_dir.mkdir(exist_ok=True)
qr_dir.mkdir(exist_ok=True)
trainingid_dir.mkdir(exist_ok=True)

BASE_URL = "https://lodigeindustries-my.sharepoint.com/:f:/r/personal/l_karuru_lodige_com/Documents/Trainings/"

training_sheets = pd.read_excel(training_list_path, sheet_name=["Manuals", "SOPs"])
training_lookup_df = pd.concat(training_sheets.values(), ignore_index=True)

def create_id_card(emp_no, emp_name, qr_file_path):
    # Paths
    photo_path = PHOTO_DIR / f"{emp_no}.png"
    qr_img = PILImage.open(qr_file_path).convert("RGB")
    print(photo_path)
    if not photo_path.exists():
        print(f"No photo for {emp_no}, skipping ID card.")
        return
    photo_img = PILImage.open(photo_path).convert("RGB")

    # Resize to a common size
    photo_img = photo_img.resize((300, 300))
    qr_img = qr_img.resize((300, 300))

    # Create blank canvas
    card = PILImage.new("RGB", (650, 400), "white")
    try:
        title_font = ImageFont.truetype("arialbd.ttf", 24)  
        footer_font = ImageFont.truetype("arial.ttf", 16)
        label_font = ImageFont.truetype("arial.ttf", 18)
    except IOError:
        title_font = ImageFont.load_default()
        footer_font = ImageFont.load_default()
        label_font = ImageFont.load_default()
    draw = ImageDraw.Draw(card)
    draw.text((220, 10), "Training ID", font=title_font, fill="black")
    # Paste images
    card.paste(photo_img, (20, 50))
    card.paste(qr_img, (330, 50))

    # Add text
    draw = ImageDraw.Draw(card)
    font = ImageFont.load_default()
    draw.text((20, 10), f"ID: {emp_no}", font=font, fill="black")
    draw.text((20, 370), f"{emp_name}", font=font, fill="black")

    draw.text((330, 425), "@copyright Lodige Systems Middle East", font=footer_font, fill="gray")

    # Save final ID card
    card.save(trainingid_dir / f"ID_CARD_{emp_no}.png")
    
def find_header_row(sheet_name, file_path):
    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
    for i, row in preview.iterrows():
        if "Emp. No." in row.values and "Employee Name" in row.values:
            return i
    raise ValueError(f"Could not find header row in sheet: {sheet_name}")

def deduplicate_columns(cols):
    seen = {}
    new_cols = []
    for col in cols:
        if col not in seen:
            seen[col] = 0
            new_cols.append(col)
        else:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
    return new_cols

training_lookup = {
    str(row["Trainings"]).strip().lower(): {
        "code": row["CODE"],
        "facility": row["FACILITY"],
        "category": row["Category"]
    }
    for _, row in training_lookup_df.iterrows()
}
def style_sheet(ws):
    # Style the first row as header (merged row should be done outside if needed)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    header_cell = ws.cell(row=1, column=1)  # First cell in first row
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.font = Font(bold=True, size=14)

    # Set alignment for the rest of the sheet
    align = Alignment(horizontal="left", vertical="center")

    # Define thin border
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = align
        cell.border = thin_border
 
    # Apply styles to all cells except the first row
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = align
            cell.border = thin_border
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):  # TRAINING DATE and EXPIRY DATE columns
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD-MMM-YYYY'
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):  # EXAM DATE column
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD-MMM-YYYY'      

    # Auto-fit column widths, skipping the first row
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[1:]:  # Skip the first row
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)  # add margin
        ws.column_dimensions[col_letter].width = adjusted_width

sheet_names = ["Cargo Trainings", "DFW Trainings", "AMH Trainings", "CBF Trainings", "SOPs", "EXAMS"]
all_dfs = {}
for sheet in sheet_names:
    header_row = find_header_row(sheet, master_path)
    df = pd.read_excel(master_path, sheet_name=sheet, header=header_row)

    cleaned_cols = []
    for col in df.columns:
        col_str = str(col)

        # Strict clean for Emp No and Employee Name
        if col_str.strip() in ["Emp. No.", "Employee Name"]:
            col_str = (
                pd.Series([col_str])
                .astype(str)
                .str.strip()
                .str.lower()
                .str.replace(r"[^\w\s]", "", regex=True)
                .iloc[0]
            )

        # ✅ Also normalize date-like headers (for training + exams to work)
        elif "date" in col_str.lower():
            col_str = (
                pd.Series([col_str])
                .astype(str)
                .str.strip()
                .str.lower()
                .str.replace(r"[^\w\s]", "", regex=True)
                .iloc[0]
            )

        cleaned_cols.append(col_str)

    df.columns = deduplicate_columns(cleaned_cols)
    all_dfs[sheet] = df

cargo_df = all_dfs[sheet_names[0]]
employees = cargo_df[['emp no', 'employee name']].drop_duplicates()
for sheet_name, sheet_df in all_dfs.items():
    if sheet_name == sheet_names[0]:  # skip Cargo (already taken)
        continue
    if 'emp no' in sheet_df.columns and 'employee name' in sheet_df.columns:
        new_emps = sheet_df[['emp no', 'employee name']].drop_duplicates()
        # Add only new employees not already in the list
        employees = pd.concat([employees, new_emps]).drop_duplicates(subset=['emp no'])

today = datetime.today()
# --- Process employees ---
for _, emp in employees.iterrows():
    emp_no = emp['emp no']
    emp_name = emp['employee name']

    # --- Collect Training Records ---
    training_records = []
    for sheet_name, df in all_dfs.items():
        if sheet_name.upper() == "EXAMS":
            continue  # exams handled separately
        rows = df[df['emp no'] == emp_no]
        for _, row in rows.iterrows():
            for date_col in [c for c in df.columns if "date" in c.lower()]:
                training_date = pd.to_datetime(row.get(date_col, None), errors="coerce")

                # ✅ Skip if no date
                if pd.isna(training_date):
                    continue

                date_index = list(df.columns).index(date_col)
                training_name = df.columns[date_index + 1] if date_index + 1 < len(df.columns) else ""

                expiry_date = training_date + timedelta(days=365 if sheet_name.lower()=="sops" else 730)
                days_left = (expiry_date - today).days if expiry_date else None
                status = (
                    "VALID" if days_left is not None and days_left >= 0
                    else "NOT VALID" if days_left is not None
                    else "NOT Applicable"
                )
                if training_name:
                    lookup_key = str(training_name).strip().lower()
                    training_info = training_lookup.get(lookup_key, {"code": "", "facility": "", "category": ""})
                    training_records.append([
                        None,
                        f"{training_name} ({sheet_name})",
                        training_info["facility"],
                        training_info["code"],
                        training_info["category"],
                        training_date.strftime('%d-%b-%Y'),
                        expiry_date.strftime('%d-%b-%Y'),
                        days_left,
                        today.strftime('%d-%b-%Y'),
                        status
                    ])

    training_df = pd.DataFrame(training_records, columns=[
        "SN", "TRAININGS","FACILITY","CODE","CATEGORY","TRAINING DATE", "EXPIRY DATE", "PERIOD TO EXPIRE", "LAST UPDATE", "STATUS"
    ])
    training_df.drop_duplicates(subset=["TRAININGS", "TRAINING DATE"], inplace=True)
    training_df["SN"] = range(1, len(training_df) + 1)

    # --- Collect Exam Records ---
    exam_records = []
    if "EXAMS" in all_dfs:
        emp_exam_rows = all_dfs["EXAMS"][all_dfs["EXAMS"]['emp no'] == emp_no]
        for _, row in emp_exam_rows.iterrows():
            for idx, col in enumerate(emp_exam_rows.columns):
                if col.startswith("date") and idx + 1 < len(emp_exam_rows.columns):
                    exam_name_col = emp_exam_rows.columns[idx + 1]
                    exam_name_display = exam_name_col.replace("_", " ").title()
                    exam_date = pd.to_datetime(row[col], errors="coerce")
                    mark = row[exam_name_col]
                    if pd.notna(mark) and str(mark).strip() != "":
                        exam_records.append([
                            None,
                            exam_name_display,
                            exam_date.strftime('%d-%b-%Y') if pd.notna(exam_date) else '',
                            mark
                        ])

    exam_df = pd.DataFrame(exam_records, columns=["SN", "EXAM", "EXAM DATE", "MARKS ATTAINED"])
    if not exam_df.empty:
        exam_df.drop_duplicates(subset=["EXAM", "EXAM DATE"], inplace=True) 
        exam_df["SN"] = range(1, len(exam_df) + 1)
        try:
            numeric_marks = exam_df["MARKS ATTAINED"].astype(str).str.replace('%', '').astype(float)
            avg = numeric_marks.mean()
            exam_df.loc["TOTAL"] = ["", "", "TOTAL AVERAGE", f"{avg:.2f}%"]
        except:
            pass

    # --- Write ONE Excel file with 2 sheets ---
    wb = Workbook()
    wb.remove(wb.active)

    ws_train = wb.create_sheet(title="Training Dashboard")
    ws_train.append([f"TRAINING DASHBOARD FOR {emp_name} ({emp_no})"])
    for r in dataframe_to_rows(training_df, index=False, header=True):
        ws_train.append(r)
    style_sheet(ws_train)

    ws_exam = wb.create_sheet(title="Exam Dashboard")
    ws_exam.append([f"EXAM DASHBOARD FOR {emp_name} ({emp_no})"])
    for r in dataframe_to_rows(exam_df, index=False, header=True):
        ws_exam.append(r)
    style_sheet(ws_exam)

    # safe_emp_name = re.sub(r'[\\/*?:"<>|]', "", str(emp_name)).strip()
    safe_emp_name = str(emp_name).strip()
    if safe_emp_name.lower() == "nan" or safe_emp_name == "" or pd.isna(emp_name):
          # If no name, just use emp_no
       filename = f"{emp_no}.xlsx"
       qr_filename = f"Qr_code_for_{emp_no}.png"

    else:
        safe_emp_name = re.sub(r'[\\/*?:"<>|]', "", safe_emp_name)  # remove invalid chars
        filename = f"{safe_emp_name} {emp_no}.xlsx"
        qr_filename = f"Qr_code_for_{safe_emp_name}_{emp_no}.png"
    file_path = output_dir/filename
    wb.save(file_path)
    
    file_url = BASE_URL + file_path.name
    qr_file_path = qr_dir/qr_filename
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(file_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(qr_file_path)
    create_id_card(emp_no, emp_name, qr_file_path)

print("Dashboards & QR codes created successfully!")
